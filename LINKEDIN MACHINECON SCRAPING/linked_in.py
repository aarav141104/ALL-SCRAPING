import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import pickle
import logging
import requests
import traceback

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df):
    df.to_excel("machinecon.xlsx", index=False)
    adjust_column_width("machinecon.xlsx")


url = "https://machinecon.aimresearch.co/"
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.get(url)
driver.implicitly_wait(15)
df = pd.DataFrame(columns=["Company Name", "Name", "Designation", "Linked In Profile"])
main_elements = driver.find_elements(
    By.XPATH,
    "//div[contains(@class,'elementor-element') and contains(@class,'elementor-element-f02a8a7') and contains(@class,'elementor-widget') and contains(@class,'elementor-widget-text-editor')]/div[contains(@class,'elementor-widget-container')]/p",
)
df_existing = pd.read_excel(
    "machinecon.xlsx",
)

# driver_2 = webdriver.Chrome(service=Service(executable_path="./chromedriver"))
# driver_2.get("https://www.linkedin.com/feed/")
# time.sleep(15)
counter = 0
for element in main_elements:
    details = {}
    details["Company Name"] = element.find_element(By.XPATH, "./strong").text
    inner_text = driver.execute_script("return arguments[0].innerText;", element)
    lines = inner_text.split("\n")
    if len(lines) >= 2:
        details["Name"] = lines[1].strip()
    if len(lines) >= 3:
        details["Designation"] = lines[2].strip()

    # details["Name"] = element.find_element(
    #     By.XPATH, "./br[1]/following-sibling::text()"
    # ).strip()

    # details["Designation"] = element.find_element(
    #     By.XPATH, "./br[2]/following-sibling::text()"
    # ).strip()

    # input_element = driver_2.find_element(
    #     By.XPATH, "//input[contains(@class,'search-global-typeahead__input')]"
    # )
    # input_element.send_keys(details["Name"] + " " + details["Company Name"])
    # input_element.send_keys(Keys.RETURN)
    # linked_in_link = driver_2.find_element(
    #     By.XPATH, f"//span[contains(text(),'{details['Name']}')]/../.."
    # ).get_attribute("href")
    # details["LinkedIn Profile"] = linked_in_link
    df.loc[counter] = details
    logging.info(f"successfully saved {counter+1} rows")
    # driver_2.get("https://www.linkedin.com/feed/")
    counter += 1

df_updated = pd.concat([df_existing, df], ignore_index=True)
df_updated.to_excel("machinecon.xlsx", sheet_name="Sheet1", index=False)
adjust_column_width("machinecon.xlsx")
