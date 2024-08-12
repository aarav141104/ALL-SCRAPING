import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import pickle
import logging
import requests
import traceback

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

df = pd.read_excel(
    "machinecon.xlsx",
)
df_sliced = df.iloc[69:]
columns_to = df_sliced[["Company Name", "Name"]]


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df):
    df.to_excel("machinecon.xlsx", index=False)
    adjust_column_width("machinecon.xlsx")


driver_2 = webdriver.Chrome(service=Service(executable_path="./chromedriver"))
driver_2.get("https://www.linkedin.com/feed/")
time.sleep(20)
counter = 69
for index, row in columns_to.iterrows():
    try:
        input_element = driver_2.find_element(
            By.XPATH, "//input[contains(@class,'search-global-typeahead__input')]"
        )
        input_element.send_keys(row["Name"] + " " + row["Company Name"])
        time.sleep(2)
        input_element.send_keys(Keys.RETURN)
        time.sleep(2)
        linked_in_link = driver_2.find_element(
            By.XPATH, f"//span[contains(text(),'{row['Name']}')]/../.."
        ).get_attribute("href")
        driver_2.back()
        time.sleep(2)
        df.loc[counter, "Linked In Profile"] = linked_in_link
        save_progress(df)
        logging.info(f"Done with {counter+1} items")
        counter += 1
    except Exception as e:
        counter += 1
        driver_2.back()
        time.sleep(2)
        continue
