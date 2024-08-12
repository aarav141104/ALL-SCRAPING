import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)


urls_to_use = pd.read_excel("links_for_square_yards.xlsx")["URLs"].tolist()
df = pd.read_excel("square_yards_first.xlsx")
df["Projects"] = None
df["Ticket Prices"] = None
options = Options()
##options.add_argument("--headless")
options.add_argument("--no-sandbox")


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


# Function to process each URL and extract the desired information
def process_link_3(url):
    logging.info(f"Starting processing URL: {url}")
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an error for bad status codes
        soup = BeautifulSoup(response.content, "html.parser")
        all_items_project = soup.find_all("div", class_="typeBox")
        all_items_project = [item.get_text() for item in all_items_project]
        ticket_prices = soup.find_all(string=re.compile(r"₹"))
        ticket_prices = [item.parent.parent.parent.get_text() for item in ticket_prices]
        ticket_prices = " | ".join(ticket_prices)
        all_items_project = " | ".join(all_items_project)
        logging.info(f"Finished processing URL: {url}")
        return all_items_project, ticket_prices
    except Exception as e:
        logging.error(f"Error processing URL {url}: {e}")
        return None


# Process URLs concurrently and save results to DataFrame
counter = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
    results = executor.map(process_link_3, urls_to_use)
    for result in results:
        if result:
            project, ticket_price = result
            df.loc[counter, "Projects"] = project
            df.loc[counter, "Ticket Prices"] = ticket_price
            logging.info(f"Row {counter} added to DataFrame")
            counter += 1

# Save the DataFrame to Excel
df.to_excel("square_yards_first.xlsx", index=False)
adjust_column_width("square_yards_first.xlsx")
logging.info("Saved results to square_yards_first.xlsx")
