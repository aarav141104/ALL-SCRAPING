import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import traceback
import random
import logging
import re
import pickle

df = pd.DataFrame(columns=["Agent Name", "Address", "Reg No", "Date of Registration"])
with open("page_sources.pkl", "rb") as file:
    page_sources = pickle.load(file)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def process_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    agent_names = soup.find_all(string=re.compile(r"Agent:"))
    agent_names = [agent_name.parent.parent for agent_name in agent_names]
    addresses = [
        agent_name.find_next_sibling().get_text() for agent_name in agent_names
    ]
    agent_names = [agent_name.get_text() for agent_name in agent_names]

    reg_nos = soup.find_all(string=re.compile(r"Reg No."))
    reg_nos = [reg_no.parent.get_text() for reg_no in reg_nos]
    registration_dates = soup.find_all(string=re.compile(r"IST 2"))
    registration_dates = [
        regist_date.parent.get_text() for regist_date in registration_dates
    ]
    return {
        "Agent Name": agent_names,
        "Address": addresses,
        "Reg No": reg_nos,
        "Date of Registration": registration_dates,
    }


page_count = 0
agnt_names = []
addresses = []
reg_nos_ = []
registration_dates = []
with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    results = executor.map(process_link, page_sources)
    for result in results:
        agnt_names.extend(result["Agent Name"])
        addresses.extend(result["Address"])
        reg_nos_.extend(result["Reg No"])
        registration_dates.extend(result["Date of Registration"])
        page_count += 1
        logging.info(f"done with {page_count} pages")


df["Agent Name"] = agnt_names
df["Address"] = addresses
df["Reg No"] = reg_nos_
if len(registration_dates) < len(df):
    registration_dates.extend([None] * (len(df) - len(registration_dates)))
df["Date of Registration"] = registration_dates
df.to_excel("goa.xlsx", index=False)
adjust_column_width("goa.xlsx")
