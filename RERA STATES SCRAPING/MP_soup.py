import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


with open("mp.pkl", "rb") as file:
    page_sources = pickle.load(file)

df = pd.DataFrame(columns=["S.No.", "Name", "Address", "Type"])


def process_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    tbody = soup.find("tbody")
    all_trs = tbody.find_all("tr")
    srnos = [tr.find_all("td")[0].get_text(strip=True) for tr in all_trs]
    name = [tr.find_all("td")[1].get_text(strip=True) for tr in all_trs]
    address = [tr.find_all("td")[2].get_text(strip=True) for tr in all_trs]
    types = [tr.find_all("td")[3].get_text(strip=True) for tr in all_trs]
    return srnos, name, address, types


all_srnos = []
all_names = []
all_addresses = []
all_types = []
counter = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    results = executor.map(process_link, page_sources)
    for result in results:
        srnos, name, address, types = result
        all_srnos.extend(srnos)
        all_names.extend(name)
        all_addresses.extend(address)
        all_types.extend(types)
        counter += 1
        logging.info(f"done with {counter}")

df["S.No."] = all_srnos
df["Name"] = all_names
df["Address"] = all_addresses
df["Type"] = all_types
df.to_excel("MP.xlsx", index=False)
adjust_column_width("MP.xlsx")
