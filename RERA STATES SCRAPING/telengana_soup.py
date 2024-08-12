import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)

df = pd.DataFrame(columns=["Sr No.", "Agent Name", "Certificate No."])
with open("telengana_page_sources.pkl", "rb") as file:
    page_sources = pickle.load(file)


def process_link(page_source):
    try:
        soup = BeautifulSoup(page_source, "html.parser")
        all_trs = soup.find("tbody").find_all("tr")
        srnos = [tr.find_all("td")[0].get_text(strip=True) for tr in all_trs]
        agent_names = [tr.find_all("td")[1].get_text(strip=True) for tr in all_trs]
        certificate_nos = [tr.find_all("td")[2].get_text(strip=True) for tr in all_trs]
        return srnos, agent_names, certificate_nos
    except:
        print("NO")


counter = 0
all_srnos = []
all_agent_names = []
all_certificate_nos = []
with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
    results = executor.map(process_link, page_sources)
    for result in results:
        srnos, agent_names, certificate_nos = result
        all_srnos.extend(srnos)
        all_agent_names.extend(agent_names)
        all_certificate_nos.extend(certificate_nos)
        counter += 1
        logging.info(f"Done with {counter} Pages of data")

df["Sr No."] = all_srnos
df["Agent Name"] = all_agent_names
df["Certificate No."] = all_certificate_nos
df.to_excel("telengana.xlsx", index=False)
adjust_column_width("telengana.xlsx")
