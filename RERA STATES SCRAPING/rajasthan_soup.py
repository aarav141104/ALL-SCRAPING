import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle

with open("rajasthan.pkl", "rb") as file:
    page_sources = pickle.load(file)

df = pd.DataFrame(
    columns=[
        "District Name",
        "Agent Name",
        "Application No. / Submission Date",
        "Registration No. / Registration Date",
    ]
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)


def process_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    tbody = soup.find_all("tbody", class_="ds4u-tbody")[1]
    trs = tbody.find_all("tr")
    district_names = [tr.find_all("td")[0].get_text(strip=True) for tr in trs]
    agent_names = [tr.find_all("td")[1].get_text(strip=True) for tr in trs]
    application_no = [tr.find_all("td")[2].get_text(strip=True) for tr in trs]
    registration_date = [tr.find_all("td")[3].get_text(strip=True) for tr in trs]
    return district_names, agent_names, application_no, registration_date


all_district_names = []
all_agent_names = []
all_application_no = []
all_registration_date = []
counter = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    results = executor.map(process_link, page_sources)
    for result in results:
        district_names, agent_names, application_no, registration_date = result
        all_agent_names.extend(agent_names)
        all_district_names.extend(district_names)
        all_application_no.extend(application_no)
        all_registration_date.extend(registration_date)
        counter += 1
        logging.info(f"Done with {counter} pages")

df["District Name"] = all_district_names
df["Agent Name"] = all_agent_names
df["Application No. / Submission Date"] = all_application_no
df["Registration No. / Registration Date"] = all_registration_date
df.to_excel("rajasthan.xlsx", index=False)
adjust_column_width("rajasthan.xlsx")


df = pd.DataFrame(
    columns=[
        "District Name",
        "Agent Name",
        "Application No. / Submission Date",
        "Registration No. / Registration Date",
    ]
)
