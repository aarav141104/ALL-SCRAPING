import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import traceback
import random
import logging
import re
import pickle


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_the_page_sources(page_sources):
    with open("page_sources.pkl", "wb") as file:
        pickle.dump(page_sources, file)
    print("Saved successfully")


options = Options()
# options.add_argument("--headless")
options.add_argument("--no-sandbox")
url = "https://rera.goa.gov.in/reraApp/#_"
driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.get(url)
WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
dropdown = driver.find_element(By.XPATH, "//select[contains(@class,'form-control')][1]")
dropdown.click()
WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
driver.find_element(By.XPATH, "//option[contains(@value,'Agent')]").click()
WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
driver.find_element(By.XPATH, "//button[contains(text(),'Search')]").click()
WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

page_sources = []
for i in range(2, 100):
    try:
        page_sources.append(driver.page_source)
        next_page_element = driver.find_element(
            By.XPATH, f"//ul[@class='pagination']/li/a[contains(text(),'{i}')]"
        )
        next_page_element.click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        logging.info(f"Done with {i} items")
    except Exception as e:
        print(f"Error occured : {e}")
        traceback.print_exc
        save_the_page_sources(page_sources)
        break
