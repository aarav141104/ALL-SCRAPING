import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)

# df = pd.read_excel("magicbricks_finally.xlsx")
df = pd.DataFrame(
    columns=[
        "Company Name",
        "About Company",
        "Deals in",
        "Name",
        "ID",
        "Operating since",
        "Properties For Sale",
        "Properties For Rent",
        "Address",
        "Operates In",
        "Project",
        "ticket_size",
        "location",
        "config",
    ]
)

with open("bangalore_magicbricks_page_sources.pkl", "rb") as file:
    magicbricks_page_sources = pickle.load(file)

ids = [tup[0] for tup in magicbricks_page_sources]
first_page_sources = [tup[1] for tup in magicbricks_page_sources]
second_page_sources = [tup[2] for tup in magicbricks_page_sources]


def process_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    try:
        dealing_in = (
            soup.find(string=re.compile(r"Dealing In"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        dealing_in = "N/A"
    try:
        operating_in = (
            soup.find(string=re.compile(r"Operating In"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        operating_in = "N/A"
    try:
        address = (
            soup.find(string=re.compile(r"Address"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        address = "N/A"
    try:
        about_the_agent = (
            soup.find(string=re.compile(r"About the Agent"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        about_the_agent = "N/A"
    try:
        agent_name = soup.find("span", class_="agntName").get_text(strip=True)
    except:
        agent_name = "N/A"
    try:
        company_name = soup.find("div", class_="agentName").get_text(strip=True)
    except:
        company_name = "N/A"
    try:
        operating_since = (
            soup.find(string=re.compile(r"Operating Since"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        operating_since = "N/A"
    try:
        properties_for_sale = (
            soup.find(string=re.compile(r"Properties for Sale"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        properties_for_sale = "N/A"
    try:
        properties_for_rent = (
            soup.find(string=re.compile(r"Properties for Rent"))
            .parent.find_next_sibling()
            .get_text(strip=True)
        )
    except:
        properties_for_rent = "N/A"

    return (
        dealing_in if len(dealing_in) > 0 else "N/A",
        operating_in if len(operating_in) > 0 else "N/A",
        address if len(address) > 0 else "N/A",
        about_the_agent if len(about_the_agent) > 0 else "N/A",
        agent_name if len(agent_name) > 0 else "N/A",
        company_name if len(company_name) > 0 else "N/A",
        operating_since if len(operating_since) > 0 else "N/A",
        properties_for_rent if len(properties_for_rent) > 0 else "N/A",
        properties_for_sale if len(properties_for_sale) > 0 else "N/A",
    )


counter = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
    results = executor.map(process_link, first_page_sources)
    for result in results:
        if result:
            (
                dealing_in,
                operating_in,
                address,
                about_the_agent,
                agent_name,
                company_name,
                operating_since,
                properties_for_rent,
                properties_for_sale,
            ) = result
            df.loc[counter, "About Company"] = about_the_agent
            df.loc[counter, "Deals in"] = dealing_in
            df.loc[counter, "Operates In"] = operating_in
            df.loc[counter, "Address"] = address
            df.loc[counter, "Company Name"] = company_name
            df.loc[counter, "Name"] = agent_name
            df.loc[counter, "Operating since"] = operating_since
            df.loc[counter, "Properties For Sale"] = properties_for_sale
            df.loc[counter, "Properties For Rent"] = properties_for_rent
            counter += 1
            logging.info(f"Added {counter} rows to the DataFrame")

df["ID"] = ids
df.to_excel("bangalore_magicbricks.xlsx", index=False)
adjust_column_width("bangalore_magicbricks.xlsx")
