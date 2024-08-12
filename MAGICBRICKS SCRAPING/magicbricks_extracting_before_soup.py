import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


page_sources_and_id = []
url = "https://www.magicbricks.com/residential-real-estate-agents-in-bangalore-pppagent"
options = Options()
# options.add_argument("--headless")
# options.add_argument("window-size=1920,1080")

driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.set_window_position(-2000, 0)
driver.get(url)
WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
driver.save_screenshot("lets_see.png")
elements = driver.find_elements(
    By.XPATH, "//span[contains(@class,'seeProDetail')]/a[1]"
)
first_page_to_next_one = [element.get_attribute("href") for element in elements]

counter = 0
while True:
    if counter != 0:
        elements = driver.find_elements(
            By.XPATH, "//span[contains(@class,'seeProDetail')]/a[1]"
        )
        first_page_to_next_one = [element.get_attribute("href") for element in elements]
    for element in first_page_to_next_one:
        try:
            details = {}
            local_driver = webdriver.Chrome(
                service=Service(executable_path="./chromedriver"), options=options
            )
            local_driver.get(element)
            WebDriverWait(local_driver, 60).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            id_ = element.split("-")[-1]
            first_page_source = local_driver.page_source
            try:
                properties_for_sale = local_driver.find_element(
                    By.XPATH,
                    "//a[contains(@class,'prop_sale_seeAll') and contains(text(),'Sale')]",
                )
                properties_for_sale.click()
                local_driver.switch_to.window(local_driver.window_handles[-1])
                WebDriverWait(local_driver, 60).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                local_driver.execute_script(
                    "window.scrollTo(0,document.body.scrollHeight);"
                )
                WebDriverWait(local_driver, 60).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                second_page_source = local_driver.page_source
            except Exception as e:
                second_page_source = "N/A"
            page_sources_and_id.append((id_, first_page_source, second_page_source))
            counter += 1
            logging.info(f"Added {counter} tuples to the list")
            local_driver.quit()
        except Exception as e:
            print(e)
            traceback.print_exc()
    try:
        next_page = driver.find_element(By.XPATH, "//a[contains(text(),'Next Page')]")
        next_page.click()
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except Exception as e:
        print("No Next Page Found")
        break

with open("bangalore_magicbricks_page_sources.pkl", "wb") as file:
    pickle.dump(page_sources_and_id, file)
