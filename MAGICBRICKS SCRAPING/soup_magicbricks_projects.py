import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
import traceback
import random
import requests
import logging
import re
import pickle


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("scraping.log"), logging.StreamHandler()],
)

df = pd.read_excel("bangalore_magicbricks.xlsx")

with open("bangalore_magicbricks_page_sources.pkl", "rb") as file:
    magicbricks_page_sources = pickle.load(file)

ids = [tup[0] for tup in magicbricks_page_sources]
first_page_sources = [tup[1] for tup in magicbricks_page_sources]
second_page_sources = [tup[2] for tup in magicbricks_page_sources]


def process_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    try:
        projects = soup.find_all("div", class_="mb-srp__card__society")
        projects = [project.get_text() for project in projects]
    except:
        projects = []
    try:
        locations = soup.find_all("h2", class_="mb-srp__card--title")
        locations = [address.get_text() for address in locations]
    except:
        locations = []
    try:
        configs = soup.find_all("div", class_="mb-srp__card__summary")
        configs = [config.get_text(separator=" , ") for config in configs]
    except:
        configs = []
    try:
        ticket_sizes = soup.find_all("div", class_="mb-srp__card__price--amount")
        ticket_sizes = [ticket_size.get_text() for ticket_size in ticket_sizes]
    except:
        ticket_sizes = []
    return (
        "|".join(projects) if len(projects) > 0 else "N/A",
        "|".join(locations) if len(locations) > 0 else "N/A",
        "|".join(configs) if len(configs) > 0 else "N/A",
        "|".join(ticket_sizes) if len(ticket_sizes) > 0 else "N/A",
    )


counter = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
    results = executor.map(process_link, second_page_sources)
    for result in results:
        if result:
            projects, locations, configs, ticket_sizes = result
            df.loc[counter, "Project"] = projects
            df.loc[counter, "ticket_size"] = ticket_sizes
            df.loc[counter, "location"] = locations
            df.loc[counter, "config"] = configs
            counter += 1
            logging.info(f"Added {counter} rows to the DataFrame")

df.to_excel("bangalore_magicbricks.xlsx", index=False)
adjust_column_width("bangalore_magicbricks.xlsx")
